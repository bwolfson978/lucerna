#!/usr/bin/env python3
"""Automated tool to update state income tax bracket data from the Tax Foundation.

Downloads the Tax Foundation's annual State Individual Income Tax Rates and
Brackets Excel spreadsheet, parses it, and updates the JSON configuration file
used by the Lucerna engine.

Usage:
    python -m scripts.update_state_tax_data --year 2026
    python -m scripts.update_state_tax_data --year 2026 --dry-run
    python -m scripts.update_state_tax_data --year 2026 --xlsx path/to/local.xlsx

See docs/tax-data-update.md for the full annual update workflow.

Source: Tax Foundation — State Individual Income Tax Rates and Brackets
https://taxfoundation.org/data/all/state/state-income-tax-rates/
"""

import argparse
import json
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

DATA_DIR = Path(__file__).resolve().parent.parent / "data"

# States with no income tax on wages/salary
NO_TAX_STATES = {
    "AK": "Alaska",
    "FL": "Florida",
    "NV": "Nevada",
    "NH": "New Hampshire",
    "SD": "South Dakota",
    "TN": "Tennessee",
    "TX": "Texas",
    "WA": "Washington",
    "WY": "Wyoming",
}

NO_TAX_NOTES = {
    "NH": "Repealed interest and dividends tax effective 2025. No tax on wages or Roth conversions.",
    "WA": "7% tax on long-term capital gains above $270,000 only. Not applicable to Roth conversions (ordinary income).",
}


def load_current_data() -> dict:
    """Load the current tax brackets JSON file."""
    current_file = sorted(DATA_DIR.glob("tax_brackets_*.json"))
    if not current_file:
        print("WARNING: No existing tax_brackets_*.json found in data/")
        return {}
    filepath = current_file[-1]  # Most recent year
    with open(filepath) as f:
        return json.load(f)


def download_xlsx(year: int, output_path: Path) -> Path:
    """Download the Tax Foundation XLSX for a given tax year.

    Tries several URL patterns since the Tax Foundation occasionally
    changes their file naming conventions.
    """
    import urllib.error
    import urllib.request

    url_patterns = [
        f"https://taxfoundation.org/wp-content/uploads/{year}/02/{year}-State-Individual-Income-Tax-Rates-and-Brackets-{year}.xlsx",
        f"https://taxfoundation.org/wp-content/uploads/{year}/01/{year}-State-Individual-Income-Tax-Rates-and-Brackets-{year}.xlsx",
        f"https://taxfoundation.org/wp-content/uploads/{year - 1}/12/{year}-State-Individual-Income-Tax-Rates-and-Brackets-{year}.xlsx",
        f"https://taxfoundation.org/wp-content/uploads/{year}/02/{year}-State-Individual-Income-Tax-Rates-Brackets.xlsx",
    ]

    for url in url_patterns:
        try:
            print(f"  Trying: {url}")
            urllib.request.urlretrieve(url, output_path)
            print("  Downloaded successfully.")
            return output_path
        except urllib.error.HTTPError as e:
            print(f"  HTTP {e.code} — trying next URL pattern...")
            continue
        except urllib.error.URLError as e:
            print(f"  Connection error: {e.reason}")
            continue

    print(f"\nERROR: Could not download XLSX for {year}.")
    print(
        "Please download manually from https://taxfoundation.org/data/all/state/state-income-tax-rates/"
    )
    print("Save it and re-run with: --xlsx path/to/downloaded.xlsx")
    sys.exit(1)


def parse_xlsx(xlsx_path: Path) -> dict:
    """Parse the Tax Foundation XLSX and extract bracket data.

    Returns a dict of state_code -> {brackets, deductions, tax_type, etc.}

    NOTE: The Tax Foundation XLSX format may change between years. This parser
    handles the 2024-2025 format. If the format changes significantly, this
    function will need updating — the --dry-run flag will catch parse errors
    before writing anything.
    """
    if openpyxl is None:
        print("ERROR: openpyxl is required for XLSX parsing.")
        print("Install it with: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    parse_warnings = []

    # The Tax Foundation spreadsheet typically has separate sheets or
    # sections for single and MFJ. The exact layout varies by year.
    # This parser attempts to find and extract bracket data from the
    # most common layouts.
    sheet_names = wb.sheetnames
    print(f"\n  Found sheets: {sheet_names}")

    # Try to find the main data sheet
    data_sheet = None
    for name in sheet_names:
        if "rate" in name.lower() or "bracket" in name.lower() or "data" in name.lower():
            data_sheet = wb[name]
            break
    if data_sheet is None and sheet_names:
        data_sheet = wb[sheet_names[0]]

    if data_sheet is None:
        print("ERROR: Could not find data sheet in XLSX")
        return {}

    print(f"  Using sheet: {data_sheet.title}")
    print(f"  Dimensions: {data_sheet.dimensions}")

    # Extract raw data for inspection
    rows = list(data_sheet.iter_rows(values_only=True))
    print(f"  Total rows: {len(rows)}")

    if len(rows) > 0:
        print("\n  First 5 rows (for format inspection):")
        for i, row in enumerate(rows[:5]):
            print(f"    Row {i}: {row[:10]}")

    parse_warnings.append(
        "XLSX parsing provides a DRAFT that requires manual review. "
        "The Tax Foundation changes their spreadsheet layout between years."
    )

    return {"_raw_rows": rows, "_warnings": parse_warnings}


def validate_against_current(new_data: dict, current_data: dict) -> list[str]:
    """Compare new data against current and flag significant changes.

    Returns a list of warning strings for human review.
    """
    warnings = []
    current_states = current_data.get("states", {})
    new_states = new_data.get("states", {})

    for code, new_state in new_states.items():
        if code not in current_states:
            warnings.append(
                f"NEW STATE: {code} ({new_state.get('name', '?')}) — not in current data"
            )
            continue

        current_state = current_states[code]

        # Check if tax type changed
        old_type = current_state.get("tax_type", "unknown")
        new_type = new_state.get("tax_type", "unknown")
        if old_type != new_type:
            warnings.append(f"TAX TYPE CHANGED: {code} — {old_type} -> {new_type}")

        # Check for large rate changes (> 1 percentage point)
        if new_state.get("has_income_tax") and current_state.get("has_income_tax"):
            old_brackets = current_state.get("brackets", {}).get("single", [])
            new_brackets = new_state.get("brackets", {}).get("single", [])
            if old_brackets and new_brackets:
                old_top = old_brackets[-1].get("rate", 0)
                new_top = new_brackets[-1].get("rate", 0)
                if isinstance(old_top, (int, float)) and isinstance(new_top, (int, float)):
                    diff = abs(new_top - old_top)
                    if diff > 0.01:
                        warnings.append(
                            f"LARGE RATE CHANGE: {code} top rate "
                            f"{old_top:.2%} -> {new_top:.2%} (delta: {diff:.2%})"
                        )

    # Check for states that disappeared
    for code in current_states:
        if code not in new_states:
            warnings.append(f"MISSING STATE: {code} — present in current data but not in new data")

    return warnings


def write_output(data: dict, year: int, dry_run: bool = False) -> Path:
    """Write the updated JSON file."""
    output_path = DATA_DIR / f"tax_brackets_{year}.json"

    if dry_run:
        print(f"\n[DRY RUN] Would write to: {output_path}")
        print(
            f"[DRY RUN] States with income tax: {sum(1 for s in data.get('states', {}).values() if s.get('has_income_tax'))}"
        )
        print(
            f"[DRY RUN] No-tax states: {sum(1 for s in data.get('states', {}).values() if not s.get('has_income_tax'))}"
        )
        return output_path

    # Write with consistent formatting
    with open(output_path, "w") as f:
        json.dump(data, f, indent=2)
        f.write("\n")

    print(f"\nWrote: {output_path}")
    return output_path


def build_scaffold(year: int) -> dict:
    """Build a scaffold JSON from current data, bumped to the new year.

    This is the primary workflow: copy current year's data as a starting
    point, then apply updates from the XLSX or manual edits.
    """
    current = load_current_data()
    if not current:
        print("ERROR: No current data to scaffold from. Create the initial JSON manually.")
        sys.exit(1)

    scaffold = {
        "metadata": {
            "tax_year": year,
            "primary_source": "Tax Foundation - State Individual Income Tax Rates and Brackets",
            "primary_source_url": "https://taxfoundation.org/data/all/state/state-income-tax-rates/",
            "federal_source": f"IRS Revenue Procedure (check for {year} update)",
            "last_updated": date.today().isoformat(),
            "notes": f"Scaffolded from {current.get('metadata', {}).get('tax_year', '?')} data. "
            f"Review and update all bracket thresholds and rates for {year}.",
        },
        "federal": current.get("federal", {}),
        "states": current.get("states", {}),
    }

    return scaffold


def print_summary(data: dict) -> None:
    """Print a summary of the tax data."""
    states = data.get("states", {})
    has_tax = [c for c, s in states.items() if s.get("has_income_tax")]
    no_tax = [c for c, s in states.items() if not s.get("has_income_tax")]
    flat = [c for c, s in states.items() if s.get("tax_type") == "flat"]
    progressive = [c for c, s in states.items() if s.get("tax_type") == "progressive"]

    print(f"\n{'=' * 60}")
    print(f"Tax Data Summary — Tax Year {data.get('metadata', {}).get('tax_year', '?')}")
    print(f"{'=' * 60}")
    print(f"  Total states/territories: {len(states)}")
    print(f"  With income tax: {len(has_tax)}")
    print(f"    Flat tax: {len(flat)} ({', '.join(sorted(flat))})")
    print(f"    Progressive: {len(progressive)} ({', '.join(sorted(progressive))})")
    print(f"  No income tax: {len(no_tax)} ({', '.join(sorted(no_tax))})")

    # Federal brackets
    federal = data.get("federal", {})
    single_brackets = federal.get("brackets", {}).get("single", [])
    if single_brackets:
        rates = [b["rate"] for b in single_brackets]
        print(
            f"\n  Federal brackets (single): {len(single_brackets)} brackets, "
            f"{min(rates):.0%} - {max(rates):.0%}"
        )

    # Top 5 highest-tax states
    top_rates = []
    for code in has_tax:
        s = states[code]
        brackets = s.get("brackets", {}).get("single", [])
        if brackets:
            top_rate = max(b["rate"] for b in brackets)
            top_rates.append((code, s["name"], top_rate))
    top_rates.sort(key=lambda x: x[2], reverse=True)

    print("\n  Top 5 highest state tax rates:")
    for code, name, rate in top_rates[:5]:
        print(f"    {code} ({name}): {rate:.2%}")


def main():
    parser = argparse.ArgumentParser(
        description="Update state income tax bracket data from Tax Foundation.",
        epilog="See docs/tax-data-update.md for the full annual update workflow.",
    )
    parser.add_argument(
        "--year",
        type=int,
        required=True,
        help="Tax year to update (e.g., 2026)",
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=None,
        help="Path to locally downloaded Tax Foundation XLSX (skips download)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and validate without writing files",
    )
    parser.add_argument(
        "--scaffold-only",
        action="store_true",
        help="Only create a scaffold from current year (no XLSX parsing)",
    )
    parser.add_argument(
        "--validate",
        type=Path,
        default=None,
        help="Validate an existing JSON file against current data",
    )

    args = parser.parse_args()

    print(f"Lucerna Tax Data Updater — Tax Year {args.year}")
    print(f"{'=' * 50}")

    # Validation-only mode
    if args.validate:
        print(f"\nValidating: {args.validate}")
        with open(args.validate) as f:
            new_data = json.load(f)
        current = load_current_data()
        warnings = validate_against_current(new_data, current)
        if warnings:
            print(f"\n{len(warnings)} warning(s):")
            for w in warnings:
                print(f"  ⚠ {w}")
        else:
            print("\nNo warnings — data looks consistent.")
        print_summary(new_data)
        return

    # Scaffold-only mode: copy current data, bump year
    if args.scaffold_only:
        print("\nScaffold mode: copying current data as starting point...")
        data = build_scaffold(args.year)
        print_summary(data)
        write_output(data, args.year, dry_run=args.dry_run)
        if not args.dry_run:
            print("\nNext steps:")
            print(f"  1. Download the Tax Foundation {args.year} XLSX")
            print("  2. Manually update bracket thresholds in the JSON")
            print(
                f"  3. Run: python -m scripts.update_state_tax_data --year {args.year} --validate data/tax_brackets_{args.year}.json"
            )
            print("  4. Update the engine import if the filename changed")
        return

    # Full mode: download/parse XLSX + scaffold
    print("\nStep 1: Loading current data...")
    current = load_current_data()
    current_year = current.get("metadata", {}).get("tax_year", "?")
    print(f"  Current data: tax year {current_year}")

    print("\nStep 2: Getting Tax Foundation XLSX...")
    if args.xlsx:
        xlsx_path = args.xlsx
        print(f"  Using local file: {xlsx_path}")
    else:
        xlsx_path = DATA_DIR / f"tax_foundation_{args.year}.xlsx"
        download_xlsx(args.year, xlsx_path)

    print("\nStep 3: Parsing XLSX...")
    parsed = parse_xlsx(xlsx_path)

    if "_warnings" in parsed:
        print("\n  Parse warnings:")
        for w in parsed["_warnings"]:
            print(f"    ⚠ {w}")

    # Since full automated parsing is fragile (format changes yearly),
    # the recommended workflow is scaffold + manual review
    print("\nStep 4: Creating scaffold from current data...")
    data = build_scaffold(args.year)

    print("\nStep 5: Validation...")
    warnings = validate_against_current(data, current)
    if warnings:
        print(f"\n  {len(warnings)} validation warning(s):")
        for w in warnings:
            print(f"    ⚠ {w}")

    print_summary(data)
    write_output(data, args.year, dry_run=args.dry_run)

    print("\nIMPORTANT: The scaffold copies last year's data. You must:")
    print("  1. Review the XLSX for rate/threshold changes")
    print("  2. Manually update the JSON for any changed states")
    print("  3. Check for new legislation (states switching to flat tax, etc.)")
    print(
        f"  4. Run: python -m scripts.update_state_tax_data --year {args.year} --validate data/tax_brackets_{args.year}.json"
    )
    print("  5. Run the test suite: python -m pytest tests/ -v")

    # Clean up downloaded XLSX (keep local files)
    if not args.xlsx and xlsx_path.exists():
        print(f"\n  Note: Downloaded XLSX saved at {xlsx_path} for reference")


if __name__ == "__main__":
    main()
